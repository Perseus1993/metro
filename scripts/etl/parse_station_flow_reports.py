"""Parse station-level passenger flow reports into a flat CSV.

Input workbooks are the daily ``线路分时段客运量统计报表`` files. The output
uses one row per date, time bin, station, and in/out direction.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))
from scripts.etl.paths import RAW_STATION_REPORTS, STAGING_STATION_FLOW_30MIN  # noqa: E402
from src.station_catalog import canonical_station_name  # noqa: E402


DATE_RE = re.compile(r"(20\d{2})[-年]?(\d{1,2})[-月]?(\d{1,2})")


def parse_date_from_text(text: str) -> str | None:
    match = DATE_RE.search(text)
    if not match:
        return None
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def parse_date_from_workbook(rows: list[tuple]) -> str | None:
    if not rows:
        return None
    for cell in rows[0]:
        if cell is None:
            continue
        date = parse_date_from_text(str(cell))
        if date:
            return date
    return None


def parse_date_from_path(path: Path) -> str:
    for text in [path.name, *reversed(path.parts)]:
        date = parse_date_from_text(text)
        if date:
            return date
    raise ValueError(f"Cannot parse report date from {path}")


def normalize_time_bin(value: object) -> tuple[str, str, str] | None:
    text = str(value).strip().replace(" ", "")
    if "-" not in text:
        return None
    start, end = text.split("-", 1)
    return text, start, end


def load_rows(path: Path) -> list[tuple]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        worksheet.reset_dimensions()
        return list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def report_files(root: Path) -> list[Path]:
    return sorted(root.rglob("*线路分时段客运量统计报表*.xlsx"))


def parse_report(path: Path) -> Iterable[dict[str, object]]:
    rows = load_rows(path)
    if len(rows) < 8:
        return

    report_date = parse_date_from_workbook(rows) or parse_date_from_path(path)
    line_row = rows[4]
    station_row = rows[5]

    column_meta: dict[int, tuple[str, str, str]] = {}
    current_line = ""
    for col_idx, (line_value, station_value) in enumerate(zip(line_row, station_row)):
        if line_value not in (None, "", "时间范围", "进出站", "合计"):
            current_line = str(line_value).strip()

        if station_value in (None, "", "合计"):
            continue

        station_display = str(station_value).strip()
        station_name = canonical_station_name(station_display)
        if not station_name:
            continue
        column_meta[col_idx] = (current_line, station_display, station_name)

    if not column_meta:
        print(f"warning=no_station_columns source_file={path}")
        return

    current_time_bin = ""
    current_start = ""
    current_end = ""
    for row in rows[6:]:
        if not row or all(value is None for value in row):
            continue

        if len(row) > 0 and row[0] not in (None, ""):
            normalized = normalize_time_bin(row[0])
            if normalized is None:
                current_time_bin = ""
                current_start = ""
                current_end = ""
                continue
            current_time_bin, current_start, current_end = normalized

        in_out = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if in_out not in ("进站", "出站"):
            continue
        if not current_time_bin:
            continue

        for col_idx, (line_label, station_display, station_name) in column_meta.items():
            value = row[col_idx] if col_idx < len(row) else None
            if value in (None, ""):
                passenger_count = 0
            else:
                passenger_count = int(float(value))

            yield {
                "date": report_date,
                "time_bin": current_time_bin,
                "start_time": current_start,
                "end_time": current_end,
                "hour": int(current_start.split(":", 1)[0]),
                "line_label": line_label,
                "station_display_name": station_display,
                "station_name": station_name,
                "in_out": in_out,
                "passenger_count": passenger_count,
                "source_file": str(path),
            }


def write_reports(input_root: Path, output_csv: Path, limit: int | None = None) -> int:
    files = report_files(input_root)
    if limit is not None:
        files = files[:limit]

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "date",
        "time_bin",
        "start_time",
        "end_time",
        "hour",
        "line_label",
        "station_display_name",
        "station_name",
        "in_out",
        "passenger_count",
        "source_file",
    ]

    rows_written = 0
    with output_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for idx, path in enumerate(files, start=1):
            file_rows = 0
            for row in parse_report(path):
                writer.writerow(row)
                rows_written += 1
                file_rows += 1
            if file_rows == 0:
                print(f"warning=no_rows_written source_file={path}")
            if idx % 25 == 0:
                print(f"parsed_files={idx}/{len(files)} rows={rows_written}")

    print(f"input_root={input_root}")
    print(f"output={output_csv}")
    print(f"files={len(files)}")
    print(f"rows={rows_written}")
    return rows_written


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input-root",
        default=str(RAW_STATION_REPORTS),
    )
    parser.add_argument(
        "--output",
        default=str(STAGING_STATION_FLOW_30MIN),
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Parse only the first N files for a quick validation run.",
    )
    args = parser.parse_args()

    write_reports(Path(args.input_root), Path(args.output), args.limit)


if __name__ == "__main__":
    main()
