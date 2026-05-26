"""Parse section-level (区间断面) flow reports into a flat CSV.

Input workbooks are the daily ``线路各区间期间分时断面流量汇总表`` files.
Each file covers one line for one day, with sheets for 上行 (up) and 下行 (down).

Structure per sheet (confirmed from data):
  Row 0: 统计时间：  2025-01-01 00:00:00~2025-01-02 00:00:00
  Row 1: 时间间隔：  15分钟
  Row 2: title row  (线路各区间期间分时断面流量汇总表)
  Row 3: header     (时间段 | 线路 | 区段 | <date> | 合计 | 工作日平均 | 周末平均)
  Row 4+: data      (time_bin | line | section | daily_flow | total | wd_avg | we_avg)

Output: one row per (date, time_bin_15min, line, direction, section, from_station, to_station).
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Iterable

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))
from scripts.etl.paths import RAW_SECTION_REPORTS, STAGING_SECTION_FLOW_15MIN  # noqa: E402
from src.station_catalog import canonical_station_name  # noqa: E402


DATE_RE = re.compile(r"(20\d{2})[-年]?(\d{1,2})[-月]?(\d{1,2})")
TIME_BIN_RE = re.compile(r"(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})")


def parse_date_from_text(text: str) -> str | None:
    match = DATE_RE.search(text)
    if not match:
        return None
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def parse_date_from_path(path: Path) -> str:
    """Fallback: extract date from directory name like '1.15' (month.day)."""
    for text in reversed(path.parts):
        date = parse_date_from_text(text)
        if date:
            return date
    # Try directory pattern like "1.15" under "1月份-..."
    for part in reversed(path.parts):
        m = re.fullmatch(r"(\d{1,2})\.(\d{1,2})", part)
        if m:
            month, day = int(m.group(1)), int(m.group(2))
            # Infer year from parent dir
            for p2 in reversed(path.parts):
                ym = re.search(r"(\d{1,2})月", p2)
                if ym:
                    return f"2025-{month:02d}-{day:02d}"
            return f"2025-{month:02d}-{day:02d}"
    raise ValueError(f"Cannot parse report date from {path}")


def parse_section(section_text: str) -> tuple[str, str] | None:
    """Parse '咸阳西站->宝泉路' into (from_station, to_station)."""
    if not section_text or "->" not in section_text:
        return None
    parts = section_text.split("->", 1)
    if len(parts) != 2:
        return None
    from_raw = parts[0].strip()
    to_raw = parts[1].strip()
    if not from_raw or not to_raw:
        return None
    return from_raw, to_raw


def normalize_time_bin(value: object) -> tuple[str, str, str] | None:
    """Parse '00:00-00:15' into (bin_str, start, end)."""
    text = str(value).strip().replace(" ", "")
    m = TIME_BIN_RE.match(text)
    if not m:
        return None
    start, end = m.group(1), m.group(2)
    return text, start, end


def direction_from_sheet(sheet_name: str) -> str:
    """Map sheet name to direction label."""
    if "上行" in sheet_name:
        return "up"
    if "下行" in sheet_name:
        return "down"
    return sheet_name


def parse_sheet(
    path: Path,
    sheet_name: str,
    rows: list[tuple],
) -> Iterable[dict[str, object]]:
    """Parse a single sheet (上行 or 下行) of a section flow workbook."""
    if len(rows) < 5:
        return

    # --- Extract date from Row 0 ---
    report_date: str | None = None
    if rows[0] and rows[0][0] is not None:
        report_date = parse_date_from_text(str(rows[0][1]) if len(rows[0]) > 1 else "")
    if report_date is None:
        report_date = parse_date_from_path(path)

    direction = direction_from_sheet(sheet_name)

    # --- Parse data rows (starting from Row 4) ---
    current_time_bin = ""
    current_start = ""
    current_end = ""

    for row in rows[4:]:
        if not row or all(v is None for v in row):
            continue

        # Column 0: time bin (may be None if continuation row)
        if row[0] is not None and str(row[0]).strip():
            normalized = normalize_time_bin(row[0])
            if normalized is None:
                current_time_bin = ""
                current_start = ""
                current_end = ""
                continue
            current_time_bin, current_start, current_end = normalized

        if not current_time_bin:
            continue

        # Column 1: line name (e.g., 地铁1号线)
        line_label = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        # Column 2: section (e.g., 咸阳西站->宝泉路)
        section_raw = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        parsed = parse_section(section_raw)
        if parsed is None:
            continue
        from_station_raw, to_station_raw = parsed
        from_station = canonical_station_name(from_station_raw)
        to_station = canonical_station_name(to_station_raw)

        # Column 3: daily flow value (the date-specific column)
        daily_flow_raw = row[3] if len(row) > 3 else None
        if daily_flow_raw is None or daily_flow_raw == "":
            daily_flow = 0
        else:
            daily_flow = int(float(daily_flow_raw))

        hour = int(current_start.split(":")[0])

        yield {
            "date": report_date,
            "time_bin_15min": current_time_bin,
            "start_time": current_start,
            "end_time": current_end,
            "hour": hour,
            "line_label": line_label,
            "direction": direction,
            "section": section_raw,
            "from_station_raw": from_station_raw,
            "to_station_raw": to_station_raw,
            "from_station": from_station,
            "to_station": to_station,
            "section_flow": daily_flow,
            "source_file": str(path),
        }


def parse_workbook(path: Path) -> Iterable[dict[str, object]]:
    """Parse all sheets in a section flow workbook."""
    from openpyxl import load_workbook
    import warnings

    warnings.filterwarnings("ignore", category=UserWarning)

    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name in ("Sheet2", "Sheet1"):
            continue
        ws = wb[sheet_name]
        ws.reset_dimensions()
        rows = list(ws.iter_rows(values_only=True))
        yield from parse_sheet(path, sheet_name, rows)
    wb.close()


def section_flow_files(root: Path) -> list[Path]:
    """Find all section flow Excel files under root, sorted."""
    return sorted(root.rglob("*线路各区间*分时断面流量汇总表*.xlsx"))


def write_reports(
    input_root: Path,
    output_csv: Path,
    limit: int | None = None,
) -> int:
    files = section_flow_files(input_root)
    if not files:
        print(f"[AUDIT] warning=no_files_found input_root={input_root}")
        return 0
    if limit is not None:
        files = files[:limit]

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "date",
        "time_bin_15min",
        "start_time",
        "end_time",
        "hour",
        "line_label",
        "direction",
        "section",
        "from_station_raw",
        "to_station_raw",
        "from_station",
        "to_station",
        "section_flow",
        "source_file",
    ]

    rows_written = 0
    files_parsed = 0
    files_empty = 0
    with output_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for idx, path in enumerate(files, start=1):
            file_rows = 0
            try:
                for row in parse_workbook(path):
                    writer.writerow(row)
                    rows_written += 1
                    file_rows += 1
            except Exception as exc:
                print(f"[AUDIT] error=parse_failed source_file={path} exc={exc}")
                continue
            if file_rows == 0:
                files_empty += 1
            files_parsed += 1
            if idx % 100 == 0:
                print(f"[AUDIT] parsed_files={idx}/{len(files)} rows={rows_written}")

    print(f"[AUDIT] input_root={input_root}")
    print(f"[AUDIT] output={output_csv}")
    print(f"[AUDIT] files_found={len(files)} files_parsed={files_parsed} files_empty={files_empty}")
    print(f"[AUDIT] rows_written={rows_written}")
    return rows_written


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Parse 线路各区间期间分时断面流量汇总表 into flat CSV."
    )
    parser.add_argument(
        "--input-root",
        default=str(RAW_SECTION_REPORTS),
    )
    parser.add_argument(
        "--output",
        default=str(STAGING_SECTION_FLOW_15MIN),
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Parse only the first N files for a quick validation run.",
    )
    args = parser.parse_args()

    write_reports(Path(args.input_root), Path(args.output), args.limit)


if __name__ == "__main__":
    main()
