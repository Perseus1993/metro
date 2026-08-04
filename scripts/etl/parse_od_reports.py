"""Parse hourly network OD reports into flat CSV outputs.

The detailed output stores only non-zero OD records. A second network-level
hourly summary is always produced because it is small and useful for quick
weather joins.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))
from scripts.etl.paths import (  # noqa: E402
    RAW_OD_REPORTS,
    STAGING_OD_NETWORK_HOURLY,
    STAGING_OD_NONZERO_60MIN,
)
from metro_data_warehouse.station_catalog import canonical_station_name  # noqa: E402


DATE_RE = re.compile(r"(20\d{2})[-年]?(\d{1,2})[-月]?(\d{1,2})")


def parse_date_from_text(text: str) -> str | None:
    match = DATE_RE.search(text)
    if not match:
        return None
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def parse_date_from_path(path: Path) -> str:
    for text in [path.name, *reversed(path.parts)]:
        date = parse_date_from_text(text)
        if date:
            return date
    raise ValueError(f"Cannot parse report date from {path}")


def normalize_time_bin(value: object) -> tuple[str, str, str] | None:
    text = str(value).strip().replace(" ", "")
    if "-" not in text:
        return None
    start, end = text.split("-", 1)
    return text, start, end


def od_files(root: Path) -> list[Path]:
    return sorted(root.rglob("*线网OD统计分析报表-分时段*.xlsx"))


def load_rows(path: Path, sheet_name: str) -> list[tuple]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        worksheet.reset_dimensions()
        return list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def destination_columns(header_row: tuple) -> dict[int, str]:
    columns = {}
    for col_idx, value in enumerate(header_row):
        if value in (None, "", "合计"):
            continue
        name = canonical_station_name(str(value).strip())
        if name:
            columns[col_idx] = name
    return columns


def parse_sheet(path: Path, sheet_name: str) -> Iterable[dict[str, object]]:
    rows = load_rows(path, sheet_name)
    if len(rows) < 8:
        return

    report_date = parse_date_from_text(str(rows[0][1])) if len(rows[0]) > 1 else None
    if report_date is None:
        report_date = parse_date_from_path(path)

    dest_cols = destination_columns(rows[6])
    if not dest_cols:
        print(f"warning=no_destination_columns sheet={sheet_name} source_file={path}")
        return

    current_time_bin = ""
    current_start = ""
    current_end = ""
    current_origin_line = ""
    seen_summary_hours: set[tuple[str, int]] = set()

    for row in rows[7:]:
        if not row or all(value is None for value in row):
            continue

        if len(row) > 0 and row[0] not in (None, ""):
            normalized = normalize_time_bin(row[0])
            if normalized is None:
                current_time_bin = ""
                current_start = ""
                current_end = ""
                continue
            current_time_bin, current_start, current_end = normalized

        if len(row) > 1 and row[1] not in (None, ""):
            current_origin_line = str(row[1]).strip()

        origin_value = row[2] if len(row) > 2 else None
        if origin_value in (None, "", "合计") or not current_time_bin:
            continue
        origin_station = canonical_station_name(str(origin_value).strip())
        if not origin_station:
            continue

        hour = int(current_start.split(":", 1)[0])
        summary_key = (report_date, hour)
        if summary_key not in seen_summary_hours:
            seen_summary_hours.add(summary_key)
            yield {
                "_summary_only": True,
                "date": report_date,
                "hour": hour,
            }

        for col_idx, destination_station in dest_cols.items():
            value = row[col_idx] if col_idx < len(row) else None
            if value in (None, ""):
                continue
            od_count = int(float(value))
            if od_count == 0:
                continue

            yield {
                "date": report_date,
                "time_bin": current_time_bin,
                "start_time": current_start,
                "end_time": current_end,
                "hour": hour,
                "origin_line": current_origin_line,
                "origin_station": origin_station,
                "destination_station": destination_station,
                "od_count": od_count,
                "source_sheet": sheet_name,
                "source_file": str(path),
            }


def parse_workbook(path: Path) -> Iterable[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()
    for sheet_name in sheet_names:
        yield from parse_sheet(path, sheet_name)


def write_reports(
    input_root: Path,
    output_csv: Path | None,
    summary_csv: Path,
    limit: int | None = None,
) -> tuple[int, int]:
    files = od_files(input_root)
    if limit is not None:
        files = files[:limit]

    if output_csv is not None:
        output_csv.parent.mkdir(parents=True, exist_ok=True)
    summary_csv.parent.mkdir(parents=True, exist_ok=True)

    detail_fields = [
        "date",
        "time_bin",
        "start_time",
        "end_time",
        "hour",
        "origin_line",
        "origin_station",
        "destination_station",
        "od_count",
        "source_sheet",
        "source_file",
    ]

    hourly_summary = defaultdict(lambda: {"od_total": 0, "nonzero_od_pairs": 0})
    detail_rows = 0
    detail_handle = None
    writer = None
    if output_csv is not None:
        detail_handle = output_csv.open("w", encoding="utf-8-sig", newline="")
        writer = csv.DictWriter(detail_handle, fieldnames=detail_fields)
        writer.writeheader()

    try:
        for idx, path in enumerate(files, start=1):
            file_rows = 0
            for row in parse_workbook(path):
                key = (row["date"], int(row["hour"]))
                if row.get("_summary_only"):
                    hourly_summary[key]
                    continue
                if writer is not None:
                    writer.writerow(row)
                hourly_summary[key]["od_total"] += int(row["od_count"])
                hourly_summary[key]["nonzero_od_pairs"] += 1
                detail_rows += 1
                file_rows += 1
            if file_rows == 0:
                print(f"warning=no_rows_written source_file={path}")
            if idx % 20 == 0:
                print(f"parsed_files={idx}/{len(files)} detail_rows={detail_rows}")
    finally:
        if detail_handle is not None:
            detail_handle.close()

    with summary_csv.open("w", encoding="utf-8-sig", newline="") as summary_handle:
        fields = ["date", "hour", "od_total", "nonzero_od_pairs"]
        writer = csv.DictWriter(summary_handle, fieldnames=fields)
        writer.writeheader()
        for (date, hour), values in sorted(hourly_summary.items()):
            writer.writerow(
                {
                    "date": date,
                    "hour": hour,
                    "od_total": values["od_total"],
                    "nonzero_od_pairs": values["nonzero_od_pairs"],
                }
            )

    print(f"input_root={input_root}")
    print(f"detail_output={output_csv or 'skipped'}")
    print(f"summary_output={summary_csv}")
    print(f"files={len(files)}")
    print(f"detail_rows={detail_rows}")
    print(f"summary_rows={len(hourly_summary)}")
    return detail_rows, len(hourly_summary)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input-root",
        default=str(RAW_OD_REPORTS),
    )
    parser.add_argument(
        "--output",
        default=str(STAGING_OD_NONZERO_60MIN),
    )
    parser.add_argument(
        "--summary-only",
        action="store_true",
        help="Do not write the detailed non-zero OD CSV.",
    )
    parser.add_argument(
        "--summary-output",
        default=str(STAGING_OD_NETWORK_HOURLY),
    )
    parser.add_argument("--limit", type=int, default=None)
    args = parser.parse_args()

    write_reports(
        Path(args.input_root),
        None if args.summary_only else Path(args.output),
        Path(args.summary_output),
        args.limit,
    )


if __name__ == "__main__":
    main()
