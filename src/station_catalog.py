"""
站名归一化 + 客运报表解析 + 站点主表构建。

数据流：
  线路分时段客运量统计报表.xlsx → 站点主表 → 匹配高德坐标 → master CSV
"""

import csv
import re
from collections import defaultdict
from pathlib import Path


# ---------------------------------------------------------------------------
# 站名别名映射
# ---------------------------------------------------------------------------
SPECIAL_NAME_ALIASES = {
    "创新港东": "创新港",
    "西工程大·西科大（临潼校区）": "西工程大·西科大",
    "西工程大·西科大(临潼校区)": "西工程大·西科大",
    "西电科大南校区·未来之瞳": "西电科大南校区",
    "西太路": "西部大道（西太路口）",
    "西部大道(西太路口)": "西部大道（西太路口）",
    "香湖湾(黄邓)": "香湖湾",
    "香湖湾（黄邓）": "香湖湾",
    "机场西(T1、T2、T3)": "机场西（T1、T2、T3）",
    "机场西(T1,T2,T3)": "机场西（T1、T2、T3）",
    "机场西(T1/T2/T3)": "机场西（T1、T2、T3）",
    "机场西（T1,T2,T3）": "机场西（T1、T2、T3）",
    "机场(T5)": "机场（T5）",
}


def canonical_station_name(name):
    """将站名归一化为标准名称。"""
    if name is None:
        return ""
    cleaned = re.sub(r"\s+", "", str(name).strip())
    if not cleaned:
        return ""

    cleaned = cleaned.replace("(地铁站)", "")
    cleaned = cleaned.replace("（地铁站）", "")
    cleaned = cleaned.replace("（黄邓）", "")
    cleaned = cleaned.replace("(黄邓)", "")
    cleaned = cleaned.replace("·未来之瞳", "")
    cleaned = re.sub(r"[（(]临潼校区[)）]", "", cleaned)
    cleaned = SPECIAL_NAME_ALIASES.get(cleaned, cleaned)
    return cleaned


def line_label_to_ref(line_label):
    if not line_label:
        return ""
    text = str(line_label).strip()
    if text == "西户线":
        return "36"
    match = re.search(r"(\d+)", text)
    return match.group(1) if match else ""


def line_sort_key(line_ref):
    if line_ref and line_ref.isdigit():
        return (0, int(line_ref))
    return (1, str(line_ref))


# ---------------------------------------------------------------------------
# 报表搜索与解析
# ---------------------------------------------------------------------------
def find_latest_passenger_report(root="data/reports"):
    """在指定目录下递归搜索最新的 '线路分时段客运量统计报表*.xlsx'。"""
    candidates = sorted(Path(root).rglob("线路分时段客运量统计报表*.xlsx"))
    if not candidates:
        return None

    def sort_key(path):
        date_key = (0, 0, 0)
        for part in reversed(path.parts):
            match = re.fullmatch(r"(\d{4})年(\d{1,2})月(\d{1,2})日", part)
            if match:
                date_key = tuple(int(v) for v in match.groups())
                break
        return date_key, path.stat().st_mtime, str(path)

    return max(candidates, key=sort_key)


def extract_report_station_entries(report_path):
    """从客运报表中提取站点条目。"""
    from openpyxl import load_workbook

    workbook = load_workbook(report_path, read_only=False, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    current_line = ""
    entries = []
    for column in range(3, worksheet.max_column + 1):
        line_label = worksheet.cell(5, column).value
        if line_label not in (None, "", "合计"):
            current_line = str(line_label).strip()

        station_name = worksheet.cell(6, column).value
        if station_name is None:
            continue

        display_name = str(station_name).strip()
        canonical_name = canonical_station_name(display_name)
        if not canonical_name or canonical_name == "合计":
            continue

        entries.append(
            {
                "name": canonical_name,
                "display_name": display_name,
                "line_label": current_line,
                "line_ref": line_label_to_ref(current_line),
                "column_index": column,
            }
        )

    return entries


def build_station_master(report_path):
    """从客运报表构建站点主表（不含坐标）。"""
    stations = {}
    for entry in extract_report_station_entries(report_path):
        station = stations.setdefault(
            entry["name"],
            {
                "name": entry["name"],
                "display_name": entry["display_name"],
                "line_labels": set(),
                "line_refs": set(),
                "report_names": set(),
            },
        )
        station["report_names"].add(entry["display_name"])
        station["line_labels"].add(entry["line_label"])
        if entry["line_ref"]:
            station["line_refs"].add(entry["line_ref"])

    station_list = []
    for station in stations.values():
        line_labels = sorted(label for label in station["line_labels"] if label)
        line_refs = sorted(station["line_refs"], key=line_sort_key)
        report_names = sorted(station["report_names"])
        station_list.append(
            {
                "name": station["name"],
                "display_name": report_names[0],
                "report_names": report_names,
                "line_labels": line_labels,
                "line_refs": line_refs,
                "primary_line_ref": line_refs[0] if line_refs else "",
            }
        )

    return sorted(
        station_list,
        key=lambda item: (
            line_sort_key(item["primary_line_ref"]),
            item["display_name"],
        ),
    )


# ---------------------------------------------------------------------------
# 坐标目录（仅高德）
# ---------------------------------------------------------------------------
def load_coordinate_catalog(amap_csv_path):
    """
    从高德站点 CSV 加载坐标目录。

    Args:
        amap_csv_path: 高德站点 CSV 路径（如 output/xian_metro_stations_amap_subway.csv）

    Returns:
        dict: canonical_name → list[dict] 坐标候选列表
    """
    catalog = defaultdict(list)
    path = Path(amap_csv_path)
    if not path.exists():
        print(f"[AUDIT] 坐标文件不存在: {path}")
        return catalog

    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            # 规则 10: CSV 列名读入后必须 strip()
            row = {k.strip(): v for k, v in row.items()}
            raw_name = row.get("name", "").strip()
            canonical_name = canonical_station_name(raw_name)
            if not canonical_name:
                continue
            try:
                lat = float(row["lat"])
                lon = float(row["lon"])
            except (KeyError, TypeError, ValueError):
                continue

            catalog[canonical_name].append(
                {
                    "name": canonical_name,
                    "source_name": raw_name,
                    "source": "amap",
                    "coord_system": "gcj02",
                    "lat": lat,
                    "lon": lon,
                }
            )
    return catalog


def attach_coordinates(station_master, coordinate_catalog):
    """
    将坐标目录中的坐标附加到站点主表。

    Returns:
        attached: list[dict] — 全部站点（含坐标或 missing）
        missing:  list[dict] — 未匹配到坐标的站点
    """
    attached = []
    missing = []

    for station in station_master:
        candidates = coordinate_catalog.get(station["name"], [])
        merged = dict(station)
        if candidates:
            best = candidates[0]
            merged.update(
                {
                    "lat": best["lat"],
                    "lon": best["lon"],
                    "coord_system": best["coord_system"],
                    "source": best["source"],
                    "source_name": best["source_name"],
                    "status": "matched",
                }
            )
        else:
            merged.update(
                {
                    "lat": None,
                    "lon": None,
                    "coord_system": "",
                    "source": "",
                    "source_name": "",
                    "status": "missing",
                }
            )
            missing.append(merged)
        attached.append(merged)

    return attached, missing


# ---------------------------------------------------------------------------
# 主表输出
# ---------------------------------------------------------------------------
def write_station_master_csv(path, station_master):
    """将站点主表写入 CSV。"""
    fieldnames = [
        "name",
        "display_name",
        "line_refs",
        "line_labels",
        "report_names",
        "lat",
        "lon",
        "coord_system",
        "source",
        "source_name",
        "status",
    ]
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for station in station_master:
            writer.writerow(
                {
                    "name": station["name"],
                    "display_name": station["display_name"],
                    "line_refs": "|".join(station["line_refs"]),
                    "line_labels": "|".join(station["line_labels"]),
                    "report_names": "|".join(station["report_names"]),
                    "lat": station["lat"] if station["lat"] is not None else "",
                    "lon": station["lon"] if station["lon"] is not None else "",
                    "coord_system": station["coord_system"],
                    "source": station["source"],
                    "source_name": station["source_name"],
                    "status": station["status"],
                }
            )
